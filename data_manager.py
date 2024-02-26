from datetime import datetime
import json
import pyshorteners
import constants

from openpyxl import load_workbook

from data_models import AirlineModel, DestinationModel, TripModel


class DataManager():

    def get_destinations_data(self, file_path='data/destinations.json'):
        with open(file_path, 'r') as file:
            data = json.load(file)
            destinations = [DestinationModel(**destination) for destination in data]
            return destinations
        
    def get_airlines_data(self, file_path='data/airlines.json'):
        with open(file_path, 'r') as file:
            data = json.load(file)
            airlines = [AirlineModel(code=code, name=name) for code, name in data.items()]
            return airlines
    
    def store_deal_details(self, deal):
        workbook = load_workbook('data/stored_deals.xlsx')
        sheet = workbook['Sheet1']
        next_row = sheet.max_row + 1

        # Writing data to the cells
        sheet.cell(row=next_row, column=1).value = constants.TODAY
        sheet.cell(row=next_row, column=2).value = deal.fly_to 
        sheet.cell(row=next_row, column=3).value = deal.price
        sheet.cell(row=next_row, column=4).value = deal.url

        workbook.save('data/stored_deals.xlsx')
    
    def shorten_url(self, long_url):
        type_tiny = pyshorteners.Shortener()
        short_url = type_tiny.tinyurl.short(long_url)
        return short_url
    
    def get_trip(self, trip):
        return TripModel(fly_from=f"{trip['cityFrom']} {trip['flyFrom']}", 
                        fly_to=f"{trip['cityTo']} {trip['flyTo']}",
                        departure_date=datetime.strptime(trip['local_departure'], '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%d/%m/%Y'),
                        length_of_stay=trip['nightsInDest'],
                        price=trip['price'],
                        airline=trip['airlines'][0],
                        url=self.shorten_url(trip['deep_link'])
                        )
    

    def prepare_params(self, destination: DestinationModel):
        params = {'fly_from': destination.fly_from,
            'fly_to': destination.fly_to,
            'date_from': destination.date_from if not None else constants.DATE_FROM,
            'date_to': destination.date_to if not None else constants.DATE_TO,
            'nights_in_dst_from': destination.stay_min,
            'nights_in_dst_to': destination.stay_max,
            'ret_from_diff_city': destination.ret_from_diff_city if not None else False,
            'ret_to_diff_city':  destination.ret_to_diff_city if not None else False,
            'max_stopovers': destination.max_stopovers if not None else 0,
            'adults': 1,
            'curr': 'PLN',
            'sort': 'price',
            'limit': 10}
        return params
